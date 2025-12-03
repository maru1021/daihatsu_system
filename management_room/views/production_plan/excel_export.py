"""
生産計画Excel出力ビュー

対象月の全ラインの生産計画データをExcelファイルに出力します。
フロントエンドで表示されているテーブルと同じ形式で出力します。
- 1KMシート: 1ライン加工 + ライン未設定加工 + 組み付け
- 2KMシート: 2ライン加工 + 組み付け
- Cシート: 鋳造
"""
from django.http import HttpResponse
from django.views import View
from management_room.auth_mixin import ManagementRoomPermissionMixin
from management_room.models import (
    DailyMachiningProductionPlan,
    DailyAssenblyProductionPlan,
    DailyMachineCastingProductionPlan,
    DailyCastingProductionPlan,
    MachiningItem,
    AssemblyItem,
    CastingItem,
    MachiningStock,
)
from manufacturing.models import MachiningLine, AssemblyLine, CastingLine, CastingMachine
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from utils.days_in_month_dates import days_in_month_dates


class ProductionPlanExcelExportView(ManagementRoomPermissionMixin, View):
    """生産計画Excel出力ビュー"""

    def get(self, request, *args, **kwargs):
        # 年月を取得
        year = int(request.GET.get('year', datetime.now().year))
        month = int(request.GET.get('month', datetime.now().month))

        # 対象月の日付リストを作成
        date_list = days_in_month_dates(year, month)

        # Excelワークブックを作成
        wb = Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除

        # 全組立ラインを取得してシートを作成
        assembly_lines = AssemblyLine.objects.filter(active=True).order_by('name')

        for assembly_line in assembly_lines:
            # シート名を作成（例: #1 -> 1KM, #2 -> 2KM）
            sheet_name = assembly_line.name.replace('#', '') + 'KM'
            self._create_assembly_sheet(wb, assembly_line, sheet_name, year, month, date_list)

        # Cシート作成（鋳造）
        self._create_casting_sheet(wb, year, month, date_list)

        # レスポンスを作成
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="production_plan_{year}_{month:02d}.xlsx"'

        wb.save(response)
        return response

    def _create_assembly_sheet(self, wb, assembly_line, sheet_name, year, month, date_list):
        """組立ラインのシートを作成（紐づく加工ライン + ライン未設定加工 + 組み付け）"""
        ws = wb.create_sheet(sheet_name)
        current_row = 1

        # この組立ラインに紐づく加工ラインを取得
        machining_lines = MachiningLine.objects.filter(assembly=assembly_line, active=True).order_by('name')

        # 各加工ラインのテーブルを出力
        for idx, machining_line in enumerate(machining_lines):
            if idx > 0:
                current_row += 1  # 1行空ける
            current_row = self._write_machining_table(ws, machining_line, month, date_list, current_row)

        # ライン未設定の加工データを取得（最初のシートにのみ出力）
        # シート名が"1KM"の場合のみ出力
        if sheet_name == '1KM':
            no_line_items = MachiningItem.objects.filter(line__isnull=True, active=True).values('name').distinct().order_by('name')
            if no_line_items.exists():
                current_row += 1  # 1行空ける
                current_row = self._write_machining_table_no_line(ws, list(no_line_items), month, date_list, current_row)

        # 組み付けデータを出力
        current_row += 1  # 1行空ける
        current_row = self._write_assembly_table(ws, assembly_line, month, date_list, current_row)

    def _create_casting_sheet(self, wb, year, month, date_list):
        """Cシートを作成（鋳造）"""
        ws = wb.create_sheet('C')
        current_row = 1

        # 全ての鋳造ラインのデータを取得
        casting_lines = CastingLine.objects.filter(active=True).order_by('name')

        for idx, casting_line in enumerate(casting_lines):
            if idx > 0:
                current_row += 1  # 1行空ける
            current_row = self._write_casting_table(ws, casting_line, month, date_list, current_row)

    def _get_machining_assembly_data(self, line, item_names, date_list):
        """加工-組付けマッピングと組付け生産計画を取得"""
        from management_room.models import AssemblyItemMachiningItemMap

        # マッピングを取得
        machining_to_assembly_map = {}
        items_obj = MachiningItem.objects.filter(line=line, name__in=item_names, active=True)
        mappings = AssemblyItemMachiningItemMap.objects.filter(
            machining_item__in=items_obj,
            active=True
        ).select_related('assembly_item', 'machining_item')

        for mapping in mappings:
            machining_name = mapping.machining_item.name
            assembly_name = mapping.assembly_item.name
            assembly_line_id = mapping.assembly_item.line_id if mapping.assembly_item.line else None
            if machining_name not in machining_to_assembly_map:
                machining_to_assembly_map[machining_name] = []
            machining_to_assembly_map[machining_name].append({
                'assembly_name': assembly_name,
                'assembly_line_id': assembly_line_id
            })

        # 組付け生産計画を取得
        assembly_item_info = [(info['assembly_name'], info['assembly_line_id'])
                              for items in machining_to_assembly_map.values()
                              for info in items]
        assembly_item_names = list(set([name for name, _ in assembly_item_info]))

        assembly_plans = DailyAssenblyProductionPlan.objects.filter(
            date__gte=date_list[0],
            date__lte=date_list[-1],
            production_item__name__in=assembly_item_names
        ).select_related('production_item', 'line')

        assembly_plans_dict = {}
        for plan in assembly_plans:
            if plan.production_item and plan.line:
                key = (plan.line.id, plan.production_item.name, plan.date, plan.shift)
                if key not in assembly_plans_dict:
                    assembly_plans_dict[key] = []
                assembly_plans_dict[key].append(plan)

        return machining_to_assembly_map, assembly_plans_dict

    def _write_machining_table(self, ws, line, month, date_list, start_row):
        """加工テーブルを書き込む（フロントエンドと同じ形式）"""
        current_row = start_row

        # タイトル行
        ws.cell(current_row, 1, f'加工ライン: {line.name}')
        ws.cell(current_row, 1).font = Font(bold=True, size=12)
        current_row += 1

        # 品番を取得
        items = MachiningItem.objects.filter(line=line, active=True).order_by('order', 'name')
        item_names = [item.name for item in items]

        if not item_names:
            return current_row

        # 加工-組付けマッピングと組付け生産計画を取得
        machining_to_assembly_map, assembly_plans_dict = self._get_machining_assembly_data(line, item_names, date_list)

        # 生産計画データを取得
        plans = DailyMachiningProductionPlan.objects.filter(
            line=line,
            date__gte=date_list[0],
            date__lte=date_list[-1]
        ).select_related('production_item').order_by('date', 'shift', 'production_item')

        plans_map = {
            (plan.date, plan.shift, plan.production_item.name): plan
            for plan in plans
            if plan.production_item and plan.shift
        }

        # 在庫データを取得
        stock_records = MachiningStock.objects.filter(
            line_name=line.name,
            date__gte=date_list[0],
            date__lte=date_list[-1],
            item_name__in=item_names
        )

        stock_map = {
            (stock.date, stock.shift, stock.item_name): stock.stock
            for stock in stock_records
            if stock.date and stock.shift and stock.item_name
        }

        # ヘッダー（共通メソッド使用）
        current_row = self._write_common_header(ws, date_list, month, plans_map, current_row)

        # 出庫数セクション
        current_row = self._write_machining_shipment_section(
            ws, item_names, date_list, current_row,
            machining_to_assembly_map, assembly_plans_dict
        )

        # 生産数セクション
        current_row = self._write_section_rows(ws, '生産数', item_names, date_list, plans_map, 'production_quantity', current_row)

        # 在庫数セクション
        current_row = self._write_machining_stock_section(ws, item_names, date_list, stock_map, current_row)

        # 残業計画セクション
        current_row = self._write_machining_overtime_section(ws, date_list, plans_map, current_row)

        # 計画停止セクション
        current_row = self._write_machining_stop_time_section(ws, date_list, plans_map, current_row)

        return current_row

    def _write_machining_table_no_line(self, ws, items, month, date_list, start_row):
        """ライン未設定の加工テーブルを書き込む"""
        current_row = start_row

        # タイトル行
        ws.cell(current_row, 1, '加工ライン: 未設定')
        ws.cell(current_row, 1).font = Font(bold=True, size=12)
        current_row += 1

        item_names = [item['name'] for item in items]

        if not item_names:
            return current_row

        # 加工-組付けマッピングと組付け生産計画を取得（ライン未設定用）
        from management_room.models import AssemblyItemMachiningItemMap
        machining_to_assembly_map = {}
        items_obj = MachiningItem.objects.filter(name__in=item_names, line__isnull=True, active=True)
        mappings = AssemblyItemMachiningItemMap.objects.filter(
            machining_item__in=items_obj,
            active=True
        ).select_related('assembly_item', 'machining_item')

        for mapping in mappings:
            machining_name = mapping.machining_item.name
            assembly_name = mapping.assembly_item.name
            assembly_line_id = mapping.assembly_item.line_id if mapping.assembly_item.line else None
            if machining_name not in machining_to_assembly_map:
                machining_to_assembly_map[machining_name] = []
            machining_to_assembly_map[machining_name].append({
                'assembly_name': assembly_name,
                'assembly_line_id': assembly_line_id
            })

        # 組付け生産計画を取得
        assembly_item_info = [(info['assembly_name'], info['assembly_line_id'])
                              for items_list in machining_to_assembly_map.values()
                              for info in items_list]
        assembly_item_names = list(set([name for name, _ in assembly_item_info]))

        assembly_plans = DailyAssenblyProductionPlan.objects.filter(
            date__gte=date_list[0],
            date__lte=date_list[-1],
            production_item__name__in=assembly_item_names
        ).select_related('production_item', 'line')

        assembly_plans_dict = {}
        for plan in assembly_plans:
            if plan.production_item and plan.line:
                key = (plan.line.id, plan.production_item.name, plan.date, plan.shift)
                if key not in assembly_plans_dict:
                    assembly_plans_dict[key] = []
                assembly_plans_dict[key].append(plan)

        # データを取得
        plans = DailyMachiningProductionPlan.objects.filter(
            production_item__name__in=item_names,
            production_item__line__isnull=True,
            date__gte=date_list[0],
            date__lte=date_list[-1]
        ).select_related('production_item').order_by('date', 'shift', 'production_item')

        plans_map = {
            (plan.date, plan.shift, plan.production_item.name): plan
            for plan in plans
            if plan.production_item and plan.shift
        }

        # ヘッダー（共通メソッド使用）
        current_row = self._write_common_header(ws, date_list, month, plans_map, current_row)

        # 出庫数セクション
        current_row = self._write_machining_shipment_section(
            ws, item_names, date_list, current_row,
            machining_to_assembly_map, assembly_plans_dict
        )

        # 生産数セクション
        current_row = self._write_section_rows(ws, '生産数', item_names, date_list, plans_map, 'production_quantity', current_row)

        return current_row

    def _write_assembly_table(self, ws, line, month, date_list, start_row):
        """組み付けテーブルを書き込む"""
        current_row = start_row

        # タイトル行
        ws.cell(current_row, 1, f'組み付けライン: {line.name}')
        ws.cell(current_row, 1).font = Font(bold=True, size=12)
        current_row += 1

        # 品番を取得
        items = AssemblyItem.objects.filter(line=line, active=True).order_by('order', 'name')
        item_names = [item.name for item in items]

        if not item_names:
            return current_row

        # データを取得
        plans = DailyAssenblyProductionPlan.objects.filter(
            line=line,
            date__gte=date_list[0],
            date__lte=date_list[-1]
        ).select_related('production_item').order_by('date', 'shift', 'production_item')

        plans_map = {
            (plan.date, plan.shift, plan.production_item.name): plan
            for plan in plans
            if plan.production_item and plan.shift
        }

        # ヘッダー（共通メソッド使用）
        current_row = self._write_common_header(ws, date_list, month, plans_map, current_row)

        # 生産数のみ
        current_row = self._write_section_rows(ws, '生産数', item_names, date_list, plans_map, 'production_quantity', current_row)

        # 残業計画セクション
        current_row = self._write_assembly_overtime_section(ws, date_list, plans_map, current_row)

        # 計画停止セクション
        current_row = self._write_assembly_stop_time_section(ws, date_list, plans_map, current_row)

        return current_row

    def _get_casting_machining_data(self, line, date_list):
        """鋳造-加工マッピングと加工生産計画を取得"""
        from management_room.models import MachiningItemCastingItemMap

        # マッピングを取得
        casting_to_machining_map = {}
        mappings = MachiningItemCastingItemMap.objects.filter(
            casting_line_name=line.name,
            active=True
        )

        for mapping in mappings:
            casting_name = mapping.casting_item_name
            if casting_name not in casting_to_machining_map:
                casting_to_machining_map[casting_name] = []
            casting_to_machining_map[casting_name].append({
                'machining_line_name': mapping.machining_line_name,
                'machining_item_name': mapping.machining_item_name
            })

        # 加工生産計画を取得
        machining_plans = DailyMachiningProductionPlan.objects.filter(
            date__gte=date_list[0],
            date__lte=date_list[-1]
        ).select_related('production_item', 'line')

        machining_plans_dict = {}
        for plan in machining_plans:
            if plan.production_item and plan.line:
                key = (plan.line.name, plan.production_item.name, plan.date, plan.shift)
                if key not in machining_plans_dict:
                    machining_plans_dict[key] = []
                machining_plans_dict[key].append(plan)

        return casting_to_machining_map, machining_plans_dict

    def _write_casting_table(self, ws, line, month, date_list, start_row):
        """鋳造テーブルを書き込む"""
        current_row = start_row

        # タイトル行
        ws.cell(current_row, 1, f'鋳造ライン: {line.name}')
        ws.cell(current_row, 1).font = Font(bold=True, size=12)
        current_row += 1

        # 品番と設備を取得
        item_names = list(CastingItem.objects.filter(line=line, active=True).order_by('name').values_list('name', flat=True).distinct())
        machines = list(CastingMachine.objects.filter(line=line, active=True).order_by('name'))

        # 品番ごとの溶湯使用量を取得
        item_molten_metal_usage = {}
        for item in CastingItem.objects.filter(line=line, active=True):
            item_molten_metal_usage[item.name] = item.molten_metal_usage or 0

        if not item_names or not machines:
            return current_row

        # 鋳造-加工マッピングと加工生産計画を取得
        casting_to_machining_map, machining_plans_dict = self._get_casting_machining_data(line, date_list)

        # データを取得
        machine_plans = DailyMachineCastingProductionPlan.objects.filter(
            line=line,
            date__gte=date_list[0],
            date__lte=date_list[-1]
        ).select_related('machine', 'production_item').order_by('date', 'shift', 'machine', 'production_item')

        # マップに変換: {(date, shift, machine_name, item_name): plan}
        plans_map = {}
        for plan in machine_plans:
            if plan.machine and plan.shift:
                if plan.production_item:
                    key = (plan.date, plan.shift, plan.machine.name, plan.production_item.name)
                else:
                    key = (plan.date, plan.shift, plan.machine.name, '')
                plans_map[key] = plan

        # 出庫データを取得
        delivery_plans = DailyCastingProductionPlan.objects.filter(
            line=line,
            date__gte=date_list[0],
            date__lte=date_list[-1]
        ).select_related('production_item').order_by('date', 'shift', 'production_item')

        delivery_map = {
            (plan.date, plan.shift, plan.production_item.name): plan
            for plan in delivery_plans
            if plan.production_item and plan.shift
        }

        # ヘッダー（共通メソッド使用）
        current_row = self._write_common_header(ws, date_list, month, plans_map, current_row,
                                                 label='設備', total_label1='日計/夜計', total_label2='合計')

        # 1. 出庫数セクション
        current_row = self._write_casting_delivery_section(
            ws, item_names, date_list, delivery_map, current_row,
            casting_to_machining_map, machining_plans_dict
        )

        # 2. 生産台数セクション（品番ごと）
        current_row = self._write_casting_production_count_section(ws, item_names, date_list, plans_map, machines, current_row)

        # 3. 在庫数セクション
        current_row = self._write_casting_inventory_section(ws, item_names, date_list, delivery_map, current_row)

        # 4. 生産計画セクション（設備ごと、品番と金型カウント表示）
        current_row = self._write_casting_production_plan_section(ws, machines, date_list, plans_map, current_row)

        # 5. 金型交換セクション
        current_row = self._write_casting_mold_change_section(ws, machines, date_list, plans_map, current_row)

        # 6. 残業計画セクション
        current_row = self._write_casting_overtime_section(ws, machines, date_list, plans_map, current_row)

        # 7. 計画停止セクション
        current_row = self._write_casting_stop_time_section(ws, machines, date_list, plans_map, current_row)

        # 8. 溶湯セクション
        current_row = self._write_casting_molten_metal_section(ws, date_list, plans_map, machines, item_molten_metal_usage, current_row)

        # 9. ポット数セクション
        current_row = self._write_casting_pot_count_section(ws, date_list, plans_map, machines, item_molten_metal_usage, current_row)

        # 10. 中子セクション
        current_row = self._write_casting_core_section(ws, item_names, date_list, plans_map, machines, current_row)

        return current_row

    def _write_section_rows(self, ws, section_name, item_names, date_list, plans_map, data_field, start_row):
        """セクション（出庫数/生産数）の行を書き込む"""
        current_row = start_row

        # 日勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                # セクション名（出庫数/生産数）
                ws.cell(current_row, 1, section_name)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                # 日勤ラベル
                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            # 各日付のデータと日勤合計を計算
            day_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                key = (date, 'day', item_name)
                if key in plans_map:
                    value = getattr(plans_map[key], data_field, None)
                    if value is not None:
                        ws.cell(current_row, col_idx, value)
                        day_total += value
                    else:
                        ws.cell(current_row, col_idx, '')
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            # 月計(直) - 日勤合計
            ws.cell(current_row, len(date_list) + 4, day_total if day_total > 0 else '')
            ws.cell(current_row, len(date_list) + 4).alignment = Alignment(horizontal='right')

            current_row += 1

        # 夜勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                # 夜勤ラベル
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            # 各日付のデータと夜勤合計を計算
            night_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                key = (date, 'night', item_name)
                if key in plans_map:
                    value = getattr(plans_map[key], data_field, None)
                    if value is not None:
                        ws.cell(current_row, col_idx, value)
                        night_total += value
                    else:
                        ws.cell(current_row, col_idx, '')
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            # 月計(直) - 夜勤合計
            ws.cell(current_row, len(date_list) + 4, night_total if night_total > 0 else '')
            ws.cell(current_row, len(date_list) + 4).alignment = Alignment(horizontal='right')

            # 月計 - 日勤+夜勤の総合計
            # 対応する日勤の行から日勤合計を取得
            day_row = current_row - len(item_names)
            day_total_value = ws.cell(day_row, len(date_list) + 4).value or 0
            combined_total = (day_total_value if isinstance(day_total_value, (int, float)) else 0) + night_total
            ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
            ws.cell(current_row, len(date_list) + 5).alignment = Alignment(horizontal='right')

            current_row += 1

        return current_row

    def _write_machining_stock_section(self, ws, item_names, date_list, stock_map, start_row):
        """加工の在庫数セクションを書き込む"""
        current_row = start_row

        # 日勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 1, '在庫数')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            for col_idx, date in enumerate(date_list, start=4):
                key = (date, 'day', item_name)
                stock_value = stock_map.get(key, None)
                ws.cell(current_row, col_idx, stock_value if stock_value is not None else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            current_row += 1

        # 夜勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            for col_idx, date in enumerate(date_list, start=4):
                key = (date, 'night', item_name)
                stock_value = stock_map.get(key, None)
                ws.cell(current_row, col_idx, stock_value if stock_value is not None else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            current_row += 1

        return current_row

    def _write_machining_overtime_section(self, ws, date_list, plans_map, start_row):
        """加工の残業計画セクションを書き込む（日勤/夜勤、品番に依存しない）"""
        current_row = start_row

        # 日勤
        ws.cell(current_row, 1, '残業計画')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(current_row, 2, '日勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから残業時間を取得（同じ日・直なら全品番で同じ値）
            overtime = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'day':
                    overtime = plans_map[key].overtime or 0
                    break

            ws.cell(current_row, col_idx, overtime if overtime > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1

        # 夜勤
        ws.cell(current_row, 2, '夜勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから残業時間を取得
            overtime = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'night':
                    overtime = plans_map[key].overtime or 0
                    break

            ws.cell(current_row, col_idx, overtime if overtime > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1
        return current_row

    def _write_machining_stop_time_section(self, ws, date_list, plans_map, start_row):
        """加工の計画停止セクションを書き込む（日勤/夜勤、品番に依存しない）"""
        current_row = start_row

        # 日勤
        ws.cell(current_row, 1, '計画停止')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(current_row, 2, '日勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから計画停止を取得（同じ日・直なら全品番で同じ値）
            stop_time = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'day':
                    stop_time = plans_map[key].stop_time or 0
                    break

            ws.cell(current_row, col_idx, stop_time if stop_time > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1

        # 夜勤
        ws.cell(current_row, 2, '夜勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから計画停止を取得
            stop_time = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'night':
                    stop_time = plans_map[key].stop_time or 0
                    break

            ws.cell(current_row, col_idx, stop_time if stop_time > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1
        return current_row

    def _write_assembly_overtime_section(self, ws, date_list, plans_map, start_row):
        """組立の残業計画セクションを書き込む（日勤/夜勤、品番に依存しない）"""
        current_row = start_row

        # 日勤
        ws.cell(current_row, 1, '残業計画')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(current_row, 2, '日勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから残業時間を取得（同じ日・直なら全品番で同じ値）
            overtime = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'day':
                    overtime = plans_map[key].overtime or 0
                    break

            ws.cell(current_row, col_idx, overtime if overtime > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1

        # 夜勤
        ws.cell(current_row, 2, '夜勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから残業時間を取得
            overtime = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'night':
                    overtime = plans_map[key].overtime or 0
                    break

            ws.cell(current_row, col_idx, overtime if overtime > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1
        return current_row

    def _write_assembly_stop_time_section(self, ws, date_list, plans_map, start_row):
        """組立の計画停止セクションを書き込む（日勤/夜勤、品番に依存しない）"""
        current_row = start_row

        # 日勤
        ws.cell(current_row, 1, '計画停止')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(current_row, 2, '日勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから計画停止を取得（同じ日・直なら全品番で同じ値）
            stop_time = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'day':
                    stop_time = plans_map[key].stop_time or 0
                    break

            ws.cell(current_row, col_idx, stop_time if stop_time > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1

        # 夜勤
        ws.cell(current_row, 2, '夜勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            # この日のどれかのプランから計画停止を取得
            stop_time = 0
            for key in plans_map:
                if key[0] == date and key[1] == 'night':
                    stop_time = plans_map[key].stop_time or 0
                    break

            ws.cell(current_row, col_idx, stop_time if stop_time > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1
        return current_row

    def _calculate_shipment_from_assembly(self, item_name, date, shift, machining_to_assembly_map, assembly_plans_dict):
        """組付け生産計画から出庫数を計算"""
        shipment = 0
        assembly_items = machining_to_assembly_map.get(item_name, [])
        for assembly_item_info in assembly_items:
            assembly_key = (
                assembly_item_info['assembly_line_id'],
                assembly_item_info['assembly_name'],
                date,
                shift
            )
            assembly_plans_list = assembly_plans_dict.get(assembly_key, [])
            for assembly_plan in assembly_plans_list:
                if assembly_plan.production_quantity:
                    shipment += assembly_plan.production_quantity
        return shipment

    def _write_machining_shipment_section(self, ws, item_names, date_list, start_row,
                                         machining_to_assembly_map, assembly_plans_dict):
        """加工の出庫数セクションを書き込む"""
        current_row = start_row

        # 日勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 1, '出庫数')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            day_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                shipment = self._calculate_shipment_from_assembly(
                    item_name, date, 'day', machining_to_assembly_map, assembly_plans_dict
                )

                if shipment > 0:
                    ws.cell(current_row, col_idx, shipment)
                    day_total += shipment
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            # 日計
            ws.cell(current_row, len(date_list) + 4, day_total if day_total > 0 else '')
            ws.cell(current_row, len(date_list) + 4).alignment = Alignment(horizontal='right')

            current_row += 1

        # 夜勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            night_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                shipment = self._calculate_shipment_from_assembly(
                    item_name, date, 'night', machining_to_assembly_map, assembly_plans_dict
                )

                if shipment > 0:
                    ws.cell(current_row, col_idx, shipment)
                    night_total += shipment
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            # 夜計
            ws.cell(current_row, len(date_list) + 4, night_total if night_total > 0 else '')
            ws.cell(current_row, len(date_list) + 4).alignment = Alignment(horizontal='right')

            # 合計（日計+夜計）
            day_row = current_row - len(item_names)
            day_total_value = ws.cell(day_row, len(date_list) + 4).value or 0
            combined_total = (day_total_value if isinstance(day_total_value, (int, float)) else 0) + night_total
            ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
            ws.cell(current_row, len(date_list) + 5).alignment = Alignment(horizontal='right')

            current_row += 1

        return current_row

    def _calculate_delivery_from_machining(self, item_name, date, shift, casting_to_machining_map, machining_plans_dict):
        """加工生産計画から出庫数を計算"""
        delivery = 0
        machining_items = casting_to_machining_map.get(item_name, [])
        for machining_item_info in machining_items:
            machining_key = (
                machining_item_info['machining_line_name'],
                machining_item_info['machining_item_name'],
                date,
                shift
            )
            machining_plans_list = machining_plans_dict.get(machining_key, [])
            for machining_plan in machining_plans_list:
                if machining_plan.production_quantity:
                    delivery += machining_plan.production_quantity
        return delivery

    def _write_casting_delivery_shift_rows(self, ws, item_names, date_list, shift, shift_label,
                                           start_row, casting_to_machining_map, machining_plans_dict,
                                           is_first_shift):
        """鋳造出庫数の直別行を書き込む"""
        current_row = start_row

        for idx, item_name in enumerate(item_names):
            if idx == 0 and is_first_shift:
                ws.cell(current_row, 1, '出庫数')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

            if idx == 0:
                ws.cell(current_row, 2, shift_label)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            shift_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                delivery = self._calculate_delivery_from_machining(
                    item_name, date, shift, casting_to_machining_map, machining_plans_dict
                )

                if delivery > 0:
                    ws.cell(current_row, col_idx, delivery)
                    shift_total += delivery
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            # 直計
            ws.cell(current_row, len(date_list) + 4, shift_total if shift_total > 0 else '')
            ws.cell(current_row, len(date_list) + 4).alignment = Alignment(horizontal='right')

            # 夜勤の場合は合計列も出力
            if shift == 'night':
                day_row = current_row - len(item_names)
                day_total_value = ws.cell(day_row, len(date_list) + 4).value or 0
                combined_total = (day_total_value if isinstance(day_total_value, (int, float)) else 0) + shift_total
                ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
                ws.cell(current_row, len(date_list) + 5).alignment = Alignment(horizontal='right')

            current_row += 1

        return current_row

    def _write_casting_delivery_section(self, ws, item_names, date_list, delivery_map, start_row,
                                        casting_to_machining_map, machining_plans_dict):
        """鋳造の出庫数セクションを書き込む"""
        current_row = start_row

        # 日勤
        current_row = self._write_casting_delivery_shift_rows(
            ws, item_names, date_list, 'day', '日勤', current_row,
            casting_to_machining_map, machining_plans_dict, is_first_shift=True
        )

        # 夜勤
        current_row = self._write_casting_delivery_shift_rows(
            ws, item_names, date_list, 'night', '夜勤', current_row,
            casting_to_machining_map, machining_plans_dict, is_first_shift=False
        )

        return current_row

    def _write_casting_production_count_section(self, ws, item_names, date_list, plans_map, machines, start_row):
        """鋳造の生産台数セクション（品番ごと、フロントエンドと同じ）を書き込む"""
        current_row = start_row

        # 日勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 1, '生産台数')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            # 各日付のデータ（全設備の合計）
            day_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                production_sum = 0
                for machine in machines:
                    key = (date, 'day', machine.name, item_name)
                    if key in plans_map:
                        production_sum += plans_map[key].production_count or 0

                ws.cell(current_row, col_idx, production_sum if production_sum > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                day_total += production_sum

            ws.cell(current_row, len(date_list) + 4, day_total if day_total > 0 else '')
            current_row += 1

        # 夜勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            # 各日付のデータ（全設備の合計）
            night_total = 0
            combined_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                production_sum = 0
                for machine in machines:
                    key = (date, 'night', machine.name, item_name)
                    if key in plans_map:
                        production_sum += plans_map[key].production_count or 0

                ws.cell(current_row, col_idx, production_sum if production_sum > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                night_total += production_sum

            # 日勤と夜勤の合計を計算
            for date in date_list:
                for machine in machines:
                    day_key = (date, 'day', machine.name, item_name)
                    night_key = (date, 'night', machine.name, item_name)
                    if day_key in plans_map:
                        combined_total += plans_map[day_key].production_count or 0
                    if night_key in plans_map:
                        combined_total += plans_map[night_key].production_count or 0

            ws.cell(current_row, len(date_list) + 4, night_total if night_total > 0 else '')
            ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
            current_row += 1

        return current_row

    def _write_casting_inventory_section(self, ws, item_names, date_list, delivery_map, start_row):
        """鋳造の在庫数セクションを書き込む"""
        current_row = start_row

        # 日勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 1, '在庫数')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            for col_idx, date in enumerate(date_list, start=4):
                key = (date, 'day', item_name)
                if key in delivery_map:
                    value = delivery_map[key].stock
                    ws.cell(current_row, col_idx, value if value is not None else '')
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            current_row += 1

        # 夜勤
        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(item_names) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, item_name)

            for col_idx, date in enumerate(date_list, start=4):
                key = (date, 'night', item_name)
                if key in delivery_map:
                    value = delivery_map[key].stock
                    ws.cell(current_row, col_idx, value if value is not None else '')
                else:
                    ws.cell(current_row, col_idx, '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            current_row += 1

        return current_row

    def _write_casting_production_plan_section(self, ws, machines, date_list, plans_map, start_row):
        """鋳造の生産計画セクション（設備ごと、品番と金型カウント表示）を書き込む"""
        current_row = start_row

        # 日勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 1, '生産計画')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(machines) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            for col_idx, date in enumerate(date_list, start=4):
                # この設備・直で生産している品番を探す
                item_name = ''
                for key in plans_map:
                    if key[0] == date and key[1] == 'day' and key[2] == machine.name:
                        plan = plans_map[key]
                        if plan.production_item:
                            item_name = plan.production_item.name
                        break

                ws.cell(current_row, col_idx, item_name)
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='center')

            current_row += 1

        # 夜勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            for col_idx, date in enumerate(date_list, start=4):
                # この設備・直で生産している品番を探す
                item_name = ''
                for key in plans_map:
                    if key[0] == date and key[1] == 'night' and key[2] == machine.name:
                        plan = plans_map[key]
                        if plan.production_item:
                            item_name = plan.production_item.name
                        break

                ws.cell(current_row, col_idx, item_name)
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='center')

            current_row += 1

        return current_row

    def _write_casting_mold_change_section(self, ws, machines, date_list, plans_map, start_row):
        """鋳造の金型交換セクションを書き込む"""
        current_row = start_row

        # 日勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 1, '金型交換')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(machines) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            day_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                mold_change = 0
                for key in plans_map:
                    if key[0] == date and key[1] == 'day' and key[2] == machine.name:
                        mold_change = plans_map[key].mold_change or 0
                        break

                ws.cell(current_row, col_idx, mold_change if mold_change > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                day_total += mold_change

            ws.cell(current_row, len(date_list) + 4, day_total if day_total > 0 else '')
            current_row += 1

        # 夜勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            night_total = 0
            combined_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                mold_change = 0
                for key in plans_map:
                    if key[0] == date and key[1] == 'night' and key[2] == machine.name:
                        mold_change = plans_map[key].mold_change or 0
                        break

                ws.cell(current_row, col_idx, mold_change if mold_change > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                night_total += mold_change

            # 日勤と夜勤の合計を計算
            for date in date_list:
                for key in plans_map:
                    if key[0] == date and key[2] == machine.name:
                        combined_total += plans_map[key].mold_change or 0

            ws.cell(current_row, len(date_list) + 4, night_total if night_total > 0 else '')
            ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
            current_row += 1

        return current_row

    def _write_casting_overtime_section(self, ws, machines, date_list, plans_map, start_row):
        """鋳造の残業計画セクションを書き込む"""
        current_row = start_row

        # 日勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 1, '残業計画')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(machines) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            day_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                overtime = 0
                for key in plans_map:
                    if key[0] == date and key[1] == 'day' and key[2] == machine.name:
                        overtime = plans_map[key].overtime or 0
                        break

                ws.cell(current_row, col_idx, overtime if overtime > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                day_total += overtime

            ws.cell(current_row, len(date_list) + 4, day_total if day_total > 0 else '')
            current_row += 1

        # 夜勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            night_total = 0
            combined_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                overtime = 0
                for key in plans_map:
                    if key[0] == date and key[1] == 'night' and key[2] == machine.name:
                        overtime = plans_map[key].overtime or 0
                        break

                ws.cell(current_row, col_idx, overtime if overtime > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                night_total += overtime

            # 日勤と夜勤の合計を計算
            for date in date_list:
                for key in plans_map:
                    if key[0] == date and key[2] == machine.name:
                        combined_total += plans_map[key].overtime or 0

            ws.cell(current_row, len(date_list) + 4, night_total if night_total > 0 else '')
            ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
            current_row += 1

        return current_row

    def _write_casting_stop_time_section(self, ws, machines, date_list, plans_map, start_row):
        """鋳造の計画停止セクションを書き込む"""
        current_row = start_row

        # 日勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 1, '計画停止')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(machines) * 2 - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(current_row, 2, '日勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            day_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                stop_time = 0
                for key in plans_map:
                    if key[0] == date and key[1] == 'day' and key[2] == machine.name:
                        stop_time = plans_map[key].stop_time or 0
                        break

                ws.cell(current_row, col_idx, stop_time if stop_time > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                day_total += stop_time

            ws.cell(current_row, len(date_list) + 4, day_total if day_total > 0 else '')
            current_row += 1

        # 夜勤
        for idx, machine in enumerate(machines):
            if idx == 0:
                ws.cell(current_row, 2, '夜勤')
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(machines) - 1, end_column=2)
                ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 3, machine.name)

            night_total = 0
            combined_total = 0
            for col_idx, date in enumerate(date_list, start=4):
                stop_time = 0
                for key in plans_map:
                    if key[0] == date and key[1] == 'night' and key[2] == machine.name:
                        stop_time = plans_map[key].stop_time or 0
                        break

                ws.cell(current_row, col_idx, stop_time if stop_time > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')
                night_total += stop_time

            # 日勤と夜勤の合計を計算
            for date in date_list:
                for key in plans_map:
                    if key[0] == date and key[2] == machine.name:
                        combined_total += plans_map[key].stop_time or 0

            ws.cell(current_row, len(date_list) + 4, night_total if night_total > 0 else '')
            ws.cell(current_row, len(date_list) + 5, combined_total if combined_total > 0 else '')
            current_row += 1

        return current_row

    def _write_casting_molten_metal_section(self, ws, date_list, plans_map, machines, item_molten_metal_usage, start_row):
        """鋳造の溶湯セクションを書き込む（生産数 × 溶湯使用量）"""
        current_row = start_row

        # 日勤
        ws.cell(current_row, 1, '溶湯')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(current_row, 2, '日勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            molten_metal_total = 0
            for machine in machines:
                for item_name, usage in item_molten_metal_usage.items():
                    key = (date, 'day', machine.name, item_name)
                    if key in plans_map:
                        production_count = plans_map[key].production_count or 0
                        molten_metal_total += production_count * usage

            ws.cell(current_row, col_idx, round(molten_metal_total) if molten_metal_total > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1

        # 夜勤
        ws.cell(current_row, 2, '夜勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            molten_metal_total = 0
            for machine in machines:
                for item_name, usage in item_molten_metal_usage.items():
                    key = (date, 'night', machine.name, item_name)
                    if key in plans_map:
                        production_count = plans_map[key].production_count or 0
                        molten_metal_total += production_count * usage

            ws.cell(current_row, col_idx, round(molten_metal_total) if molten_metal_total > 0 else '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1
        return current_row

    def _write_casting_pot_count_section(self, ws, date_list, plans_map, machines, item_molten_metal_usage, start_row):
        """鋳造のポット数セクションを書き込む（溶湯 / 1200 を小数点第1位で切り上げ）"""
        current_row = start_row

        # 日勤
        ws.cell(current_row, 1, 'ポット数')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(current_row, 2, '日勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            molten_metal_total = 0
            for machine in machines:
                for item_name, usage in item_molten_metal_usage.items():
                    key = (date, 'day', machine.name, item_name)
                    if key in plans_map:
                        production_count = plans_map[key].production_count or 0
                        molten_metal_total += production_count * usage

            if molten_metal_total > 0:
                import math
                pot_count = math.ceil(molten_metal_total / 1200 * 10) / 10
                ws.cell(current_row, col_idx, f"{pot_count:.1f}")
            else:
                ws.cell(current_row, col_idx, '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1

        # 夜勤
        ws.cell(current_row, 2, '夜勤')
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, date in enumerate(date_list, start=4):
            molten_metal_total = 0
            for machine in machines:
                for item_name, usage in item_molten_metal_usage.items():
                    key = (date, 'night', machine.name, item_name)
                    if key in plans_map:
                        production_count = plans_map[key].production_count or 0
                        molten_metal_total += production_count * usage

            if molten_metal_total > 0:
                import math
                pot_count = math.ceil(molten_metal_total / 1200 * 10) / 10
                ws.cell(current_row, col_idx, f"{pot_count:.1f}")
            else:
                ws.cell(current_row, col_idx, '')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

        current_row += 1
        return current_row

    def _write_casting_core_section(self, ws, item_names, date_list, plans_map, machines, start_row):
        """鋳造の中子セクションを書き込む（品番ごとの生産数）"""
        current_row = start_row

        for idx, item_name in enumerate(item_names):
            if idx == 0:
                ws.cell(current_row, 1, '中子')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(item_names) - 1, end_column=1)
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(current_row, 2, item_name)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')

            for col_idx, date in enumerate(date_list, start=4):
                # 全設備のこの品番の生産数を合計（日勤+夜勤）
                production_sum = 0
                for machine in machines:
                    day_key = (date, 'day', machine.name, item_name)
                    night_key = (date, 'night', machine.name, item_name)
                    if day_key in plans_map:
                        production_sum += plans_map[day_key].production_count or 0
                    if night_key in plans_map:
                        production_sum += plans_map[night_key].production_count or 0

                # 中子は24の倍数に丸める
                if production_sum > 0:
                    core_count = round(production_sum / 24) * 24
                else:
                    core_count = 0

                ws.cell(current_row, col_idx, core_count if core_count > 0 else '')
                ws.cell(current_row, col_idx).alignment = Alignment(horizontal='right')

            current_row += 1

        return current_row

    def _get_work_status_flag(self, date, plans_map):
        """休出・定時フラグを取得"""
        is_regular = False
        has_data = False

        for key in plans_map:
            if key[0] == date:
                has_data = True
                is_regular = plans_map[key].regular_working_hours
                break

        is_weekend = date.weekday() >= 5
        if is_regular:
            return '定時'
        elif is_weekend and has_data:
            return '休出'
        else:
            return ''

    def _write_common_header(self, ws, date_list, month, plans_map, start_row, label='完成品番', total_label1='月計(直)', total_label2='月計'):
        """共通のヘッダー（休出/定時、日付、稼働率）を書き込む"""
        current_row = start_row

        # ヘッダー行1: 定時・休出フラグ
        ws.cell(current_row, 1, label)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=3)

        for col_idx, date in enumerate(date_list, start=4):
            flag = self._get_work_status_flag(date, plans_map)
            ws.cell(current_row, col_idx, flag)
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='center')

        ws.cell(current_row, len(date_list) + 4, total_label1)
        ws.cell(current_row, len(date_list) + 5, total_label2)

        self._apply_header_style(ws, current_row, 1, len(date_list) + 5)
        current_row += 1

        # ヘッダー行2: 日付と曜日
        for col_idx, date in enumerate(date_list, start=4):
            weekday_names = ['月', '火', '水', '木', '金', '土', '日']
            ws.cell(current_row, col_idx, f'{month}/{date.day}\n({weekday_names[date.weekday()]})')
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal='center', wrap_text=True)

        self._apply_header_style(ws, current_row, 4, len(date_list) + 3)
        current_row += 1

        # ヘッダー行3: 稼働率（省略）
        current_row += 1

        return current_row

    def _apply_header_style(self, ws, row, start_col, end_col):
        """ヘッダー行にスタイルを適用"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='81c3f9', end_color='81c3f9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
