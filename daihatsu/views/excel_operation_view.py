import os
import pandas as pd
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.views import View
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from daihatsu.except_output import except_output


# Excel操作を行うViewの基底クラス
class ExcelOperationView(View):
    """Excel操作を行うViewの基底クラス"""
    export_model = None
    import_model = None
    excel_file_name = None
    table_class = None

    # Excelの出力形式の定義
    def set_excel_file(self, models):
        pass

    def set_select(self, ws, select_list, total_row_count):
        select_row_count = total_row_count + 300
        for select in select_list:
            dv_column = DataValidation(type="list", formula1=f'"{",".join(select["select_list"])}"', allow_blank=True)
            dv_column.add(f'{select["column"]}2:{select["column"]}{select_row_count}')
            ws.add_data_validation(dv_column)

    # 列幅を調整する
    def adjust_column_width(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column[:5]:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            # 日本語文字の場合は幅を調整（1文字 = 約2.5文字分）
            adjusted_width = max_length * 2.5
            # 最小幅と最大幅を設定
            adjusted_width = max(10, min(adjusted_width, 50))

            ws.column_dimensions[column_letter].width = adjusted_width

    # 追加の選択肢を追加する
    def extra_select(self, select_list):
        pass

    def get(self, request, *args, **kwargs):
        from openpyxl.utils import get_column_letter

        try:
            models = self.export_model

            data_list = self.set_excel_file(models)

            df = pd.DataFrame(data_list)
            total_row_count = len(df)

            # Excelファイルに出力
            excel_file = self.excel_file_name
            df.to_excel(excel_file, index=False)

            # Excelファイルを開いてデータ検証を追加
            wb = load_workbook(excel_file)
            ws = wb.active

            select_list = [{
                'column': 'A',
                'select_list': ['追加', '編集', '削除']
            }]

            # アクティブの列が必ず右端にあると仮定
            if df.columns[-1] == "アクティブ":
                last_column_index = len(df.columns)
                last_column_letter = get_column_letter(last_column_index)
                select_list.append({
                    'column': last_column_letter,
                    'select_list': ['有効', '無効']
                })

            # select_listに選択肢を追加する
            self.extra_select(select_list)

            # 操作列、アクティブ列を追加
            self.set_select(ws, select_list, total_row_count)

            # 列幅を調整
            self.adjust_column_width(ws)

            # ファイルを保存
            wb.save(excel_file)

            # ファイルをレスポンスとして返す
            with open(excel_file, 'rb') as file:
                response = HttpResponse(
                    file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{excel_file}"'

            # 一時ファイルを削除
            if os.path.exists(excel_file):
                os.remove(excel_file)

            return response

        except Exception as e:
            except_output('Export excel error', e)
            raise Exception(e)

    def validate_data(self, row, operation):
        pass

    def model_create(self, create_list):
        pass

    def model_update(self, update_list, existing_objects):
        pass

    def model_delete(self, delete_list):
        try:
            # IDリストを取得
            id_list = [row.get('ID') for row in delete_list if row.get('ID')]

            if id_list:
                deleted_count, _ = self.import_model.objects.filter(id__in=id_list).delete()
                return deleted_count
            return 0
        except Exception as e:
            except_output('Model delete error', e)
            raise Exception(e)

    def get_expected_columns(self):
        try:
            # サンプルデータを取得して列名を抽出(Excelファイルアップロード時のバリデーションに使用)
            sample_data = self.set_excel_file(self.export_model[:1])
            return list(sample_data[0].keys())
        except Exception as e:
            except_output('Get expected columns error', e)

    def post(self, request, *args, **kwargs):
        try:
            # アップロードされたファイルを取得
            uploaded_file = request.FILES.get('excel_file')
            if not uploaded_file:
                return JsonResponse({'status': 'error', 'message': 'ファイルがアップロードされていません。'})

            # ファイルサイズ制限（10MB）
            max_size = 10 * 1024 * 1024  # 10MB
            if uploaded_file.size > max_size:
                return JsonResponse({'status': 'error', 'message': 'ファイルサイズが10MBを超えています。'})

            # ファイル拡張子の検証
            allowed_extensions = ['.xlsx', '.xls']
            file_extension = uploaded_file.name.lower().split('.')[-1] if '.' in uploaded_file.name else ''
            if f'.{file_extension}' not in allowed_extensions:
                return JsonResponse({'status': 'error', 'message': 'Excelファイル（.xlsx, .xls）のみアップロード可能です。'})

            # Content-Typeの検証
            allowed_content_types = [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel',  # .xls
                'application/octet-stream'  # 一部ブラウザでの.xlsxファイル
            ]
            if uploaded_file.content_type not in allowed_content_types:
                return JsonResponse({'status': 'error', 'message': '不正なファイル形式です。'})

            # Excelファイルを読み込み(dtypeで指定しないとID列がfloatとして読み込まれる)
            try:
                df = pd.read_excel(uploaded_file, dtype={'ID': str})
            except Exception as excel_error:
                except_output('Excel file read error', excel_error)
                return JsonResponse({'status': 'error', 'message': 'Excelファイルの読み込みに失敗しました。ファイルが破損している可能性があります。'})

            # 列名バリデーション(誤ったExcelファイルをアップロード防止)
            expected_columns = self.get_expected_columns()
            input_columns = list(df.columns)

            if expected_columns != input_columns:
                return JsonResponse({
                    'status': 'error',
                    'message': 'インポートするExcelファイルが誤っています。'
                })

            # NaN値を適切に処理
            df = df.fillna('')

            # トランザクション内で処理を実行
            with transaction.atomic():
                create_list = []
                update_list = []
                delete_list = []
                results = []

                invalid_ids = [id for id in df.get('ID') if id and not str(id).strip().isdigit()]
                id_list = [int(id) for id in df.get('ID') if id and str(id).strip().isdigit()]
                update_models = self.import_model.objects.filter(id__in=id_list)
                update_models_dict = {model.id: model for model in update_models}

                # バリデーションを行ってから操作ごとにまとめる
                for index, row in df.iterrows():
                    operation = row.get('操作')
                    index += 2

                    if operation == '編集':
                        id = int(row.get('ID')) if row.get('ID').isdigit() else row.get('ID')
                        if not id:
                            results.append(f'{index}行目: IDが指定されていません。')
                            continue
                        elif id in id_list:
                            model = update_models_dict.get(id)
                            if not model:
                                results.append(f'{index}行目: ID:{id}のデータが見つかりません。')
                                continue
                            else:
                                error = self.validate_data(index, row, id)
                                if error:
                                    results.append(f'{operation}失敗: {error}')
                                    continue
                                else:
                                    update_list.append(row)
                        elif id in invalid_ids:
                            results.append(f'{index}行目: idは整数にして下さい。')
                            continue

                    elif operation == '追加':
                        error = self.validate_data(index, row, None)
                        if error:
                            results.append(f'{operation}失敗: {error}')
                            continue
                        else:
                            create_list.append(row)

                    elif operation == '削除':
                        delete_list.append(row)

                user = request.user

                # 操作ごとにまとめたものを実行する
                if create_list:
                    try:
                        success_count = self.model_create(create_list, user)
                        results.append(f'追加成功: {success_count}件')
                    except Exception as e:
                        results.append(f'追加失敗: {str(e)}')

                if update_list:
                    try:
                        success_count = self.model_update(update_list, update_models_dict, user)
                        results.append(f'編集成功: {success_count}件')
                    except Exception as e:
                        results.append(f'編集失敗: {str(e)}')

                if delete_list:
                    try:
                        success_count = self.model_delete(delete_list)
                        results.append(f'削除成功: {success_count}件')
                    except Exception as e:
                        results.append(f'削除失敗: {str(e)}')

            # データベースの変更を確実にコミット
            transaction.commit()

            # table_classのインスタンスを作成してget_context_dataを取得
            table_instance = self.table_class()
            table_instance.request = request
            context = table_instance.get_context_data()

            return JsonResponse({
                'status': 'success',
                'message': 'Excelファイルの処理が完了しました。',
                'results': results,
                'table_data': context.get('data', [])
            })

        except Exception as e:
            except_output('Import excel error', e)
            return JsonResponse({'status': 'error', 'message': 'ファイルの処理中にエラーが発生しました。'})
